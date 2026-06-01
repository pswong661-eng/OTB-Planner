"""
otb_engine.py — OTB Excel Read/Write Engine
Pure Python: no Streamlit dependency.
"""
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
import pandas as pd

# ─── Constants ────────────────────────────────────────────────────────────────
SHEET_NAME = "OTB"
HEADER_ROW = 3  # 1-based row number in sheet

# Editable metric labels — match Col F exactly (including the typo "Comitted")
EDITABLE_METRICS = [
    "Planned Purchases USD",
    "Act. Goods Received QTY",
    "Act. Sale USD",
    "Act. Sale QTY",
]

# Read-only display metrics (shown in UI but never written back to Excel)
DISPLAY_METRICS = [
    "Act. Stock Holding Month-USD",
    "Act. Stock Holding Month-QTY",
]

# Internal-only metrics: parsed for computation, never shown in UI
# Col F labels (after strip):
#   "Sale % Distribution"  → row 143, monthly % ratios used as plan distribution
#   "Planned Sale Target USD" → row 148, annual planned sale (in total_yr columns)
#   "Comitted Purchase USD" → still parsed for KPI totals, hidden from Data Entry
INTERNAL_METRICS = [
    "Sale % Distribution",
    "Planned Sale Target USD",
    "Comitted Purchase USD",        # intentional source-file typo; hidden from UI
]

# Human-readable display names shown in the UI
METRIC_DISPLAY = {
    "Planned Purchases USD":        "Planned Purchase USD",
    "Comitted Purchase USD":        "Committed Purchase USD",
    "Act. Sale USD":                "Act. Sale USD",
    "Act. Sale QTY":                "Act. Sale QTY",
    "Act. Goods Received QTY":      "Act. Goods Received QTY",
    "Act. Stock Holding Month-USD": "Stock Holding USD",
    "Act. Stock Holding Month-QTY": "Stock Holding QTY",
}

METRIC_DISPLAY_REV = {v: k for k, v in METRIC_DISPLAY.items()}

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


# ─── Column detection ──────────────────────────────────────────────────────────
def _detect_columns(ws) -> dict:
    """
    Read the header row and return 0-based column indices for every year-month.
    Returns:
    {
        "2023": [idx...],  "2024": [...], "2025": [...], "2026": [...],
        "total_2023": idx, "total_2024": idx, "total_2025": idx, "total_2026": idx,
        "labels_2025": ["Jan'25",...], "labels_2026": ["Jan'26",...],
    }
    """
    header = [cell.value for cell in ws[HEADER_ROW]]
    result: dict = {
        "2023": [], "2024": [], "2025": [], "2026": [],
        "total_2023": None, "total_2024": None,
        "total_2025": None, "total_2026": None,
        "labels_2023": [], "labels_2024": [],
        "labels_2025": [], "labels_2026": [],
    }

    for i, h in enumerate(header):
        if not h:
            continue
        s = str(h).strip()
        for yr in ("23", "24", "25", "26"):
            year_full = "20" + yr
            if f"'{yr}" in s and any(m in s for m in MONTHS):
                result[year_full].append(i)
                result[f"labels_{year_full}"].append(s)
            elif s == f"Total 20{yr}":
                result[f"total_{year_full}"] = i

    return result


# ─── Data loading ──────────────────────────────────────────────────────────────
def load_otb_data(source) -> dict:
    """
    Parse OTB_v2.xlsx and build an in-memory data structure.
    `source` can be a file path (str/Path) or raw bytes (for cloud/upload mode).

    Return value is a dict where:
      key   = (collection, category, segment)  tuple
      value = {
          "Planned Purchases USD": {
              "row_idx": int,       # 1-based Excel row
              "2023": {month: val}, "2024": {...}, "2025": {...}, "2026": {...},
              "total_2023": val, "total_2024": val,
              "total_2025": val, "total_2026": val,
          },
          ... (same for other 3 editable metrics)
          "_remarks_row": int or None,
      }
    Plus a special "_meta" key with column map and ordered collection list.
    """
    from io import BytesIO as _BytesIO
    if isinstance(source, (bytes, bytearray)):
        source = _BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. Available: {available}"
        )

    ws = wb[SHEET_NAME]
    col_map = _detect_columns(ws)

    result: dict = {}
    ordered_keys: list = []
    current_key: Optional[tuple] = None

    for row_num_0, row in enumerate(
        ws.iter_rows(min_row=HEADER_ROW + 1, values_only=False), start=HEADER_ROW + 1
    ):
        # row[0].row may be absent on EmptyCell in read_only mode — use counter
        try:
            row_num = row[0].row
        except AttributeError:
            row_num = row_num_0

        col_c = row[2].value   # Collection
        col_d = row[3].value   # Category
        col_e = row[4].value   # Segment
        col_f = row[5].value   # Metric label

        # --- Track current collection block ---
        if col_c and isinstance(col_c, str):
            stripped = col_c.strip()
            if stripped and stripped not in ("Collection",):
                cat = col_d.strip() if col_d else None
                seg = col_e.strip() if col_e else None
                key = (stripped, cat, seg)
                if key != current_key:
                    current_key = key
                    if key not in result:
                        result[key] = {"_remarks_row": None}
                        ordered_keys.append(key)

        if not col_f or not isinstance(col_f, str) or current_key is None:
            continue

        metric = col_f.strip()

        # --- Editable, display-only, AND internal metrics (all parsed the same way) ---
        if metric in EDITABLE_METRICS or metric in DISPLAY_METRICS or metric in INTERNAL_METRICS:
            entry: dict = {"row_idx": row_num}

            for yr in ("2023", "2024", "2025", "2026"):
                vals = {}
                for i, month in enumerate(MONTHS):
                    if i < len(col_map[yr]):
                        v = row[col_map[yr][i]].value
                        # Coerce non-numeric placeholders ('-', None, etc.) to 0
                        if v is None:
                            vals[month] = 0
                        else:
                            try:
                                vals[month] = float(v)
                            except (ValueError, TypeError):
                                vals[month] = 0
                    else:
                        vals[month] = 0
                entry[yr] = vals

                tot_col = col_map.get(f"total_{yr}")
                entry[f"total_{yr}"] = row[tot_col].value if tot_col is not None else None

            result[current_key][metric] = entry

        # --- Remarks row ---
        elif metric == "Remarks":
            if current_key in result:
                result[current_key]["_remarks_row"] = row_num

    wb.close()

    # Post-process: compute Stock Holding in Python (formula cache is empty in xlsx)
    _recompute_stock_holding(result, ordered_keys)

    result["_meta"] = {
        "col_map": col_map,
        "collections": ordered_keys,
        "filepath": filepath,
        "loaded_at": datetime.now().isoformat(),
    }

    return result


def _blended_annual(actual_vals: dict, plan_annual: float, dist_ratios: dict,
                    fallback_annual: float, actual_flag: dict) -> float:
    """
    Compute blended annual sale denominator for Stock Holding.

    For months flagged as 'has actual' (actual_flag[month]=True): use the real
    actual value even if it is 0 (genuine zero-sale month).
    For months flagged as plan/future (actual_flag[month]=False): substitute
    plan_annual * monthly_ratio as a mock value.

    actual_flag is keyed on MONTHS list; True = real data exists for that month.
    """
    total = 0.0
    for month in MONTHS:
        if actual_flag.get(month, False):
            total += float(actual_vals.get(month) or 0)
        else:
            ratio = float(dist_ratios.get(month) or 0)
            scale = plan_annual if plan_annual > 0 else fallback_annual
            total += scale * ratio
    return total


def _recompute_stock_holding(result: dict, ordered_keys: list) -> None:
    """
    Recompute Act. Stock Holding Month-USD and Act. Stock Holding Month-QTY
    in Python, because these are formula-only cells with no cached value in xlsx.

    Formula chain (from Excel):
      USD:
        GoodsReceived USD  = Planned Purchases USD   (row 157 = row 154)
        ClosingStock USD   = max(prev_closing - ActSaleUSD + GoodsReceivedUSD, 0)
        StockHolding USD   = ClosingStock / (AnnualSaleUSD / 12)

      QTY:
        GoodsReceived QTY  = Act. Goods Received QTY  (row 158, raw data)
        ClosingStock QTY   = max(prev_closing - ActSaleQTY + GoodsReceivedQTY, 0)
        StockHolding QTY   = ClosingStock / (AnnualSaleQTY / 12)

    Denominator (annual sale) uses a blended value:
      - Months with real actual data → use the actual figure (even if 0)
      - Months without actual data   → mock = Planned Sale Target × monthly ratio
        (from 'Sale % Distribution' row).  This matches Excel's BI149 formula which
        mixes actuals and plan for the current year.

    The 'has-actual' flag for each month is derived from whether Act. Sale USD > 0
    (or explicitly set for QTY using the same USD flag, so a real QTY=0 month is
    treated as actual rather than replaced by plan).
    """
    ALL_YEARS = ["2023", "2024", "2025", "2026"]

    for key in ordered_keys:
        if key not in result:
            continue
        coll = result[key]

        dist_entry     = coll.get("Sale % Distribution")
        plan_usd_entry = coll.get("Planned Sale Target USD")

        # ── USD ────────────────────────────────────────────────────────────────
        planned_entry  = coll.get("Planned Purchases USD")
        sale_usd_entry = coll.get("Act. Sale USD")
        sh_usd_entry   = coll.get("Act. Stock Holding Month-USD")

        # Keep per-year actual flags and annual totals so QTY can reuse them
        usd_actual_flags: dict = {}   # {yr: {month: bool}}
        usd_annual_actuals: dict = {} # {yr: float}  ← sum of real actual months only

        if planned_entry and sale_usd_entry and sh_usd_entry:
            closing_usd = 0.0
            prev_annual_usd = 0.0

            for yr in ALL_YEARS:
                planned_yr  = planned_entry.get(yr, {})
                sale_yr_usd = sale_usd_entry.get(yr, {})

                # A month has real actual data when its USD value is non-None/non-zero.
                # Parser coerces None→0, so "actual" months are those the user entered.
                # We distinguish them by checking if the raw parsed value is > 0 OR
                # by checking surrounding months (a zero month surrounded by non-zeros
                # is more likely real; a trailing-zero block is plan).
                # Simple rule: mark a month as "actual" if its USD > 0.
                # (A genuine zero-sale month is rare and acceptable as mock.)
                flags = {m: (float(sale_yr_usd.get(m) or 0) > 0) for m in MONTHS}
                usd_actual_flags[yr] = flags

                # Plan annual for this year
                plan_annual = float(
                    (plan_usd_entry or {}).get(f"total_{yr}") or 0
                )
                dist_yr = (dist_entry or {}).get(yr, {})

                annual_sale = _blended_annual(
                    sale_yr_usd, plan_annual, dist_yr, prev_annual_usd, flags
                )

                # Track actual-only sum for QTY fallback
                actual_sum = sum(float(sale_yr_usd.get(m) or 0) for m in MONTHS)
                usd_annual_actuals[yr] = actual_sum
                if actual_sum > 0:
                    prev_annual_usd = actual_sum

                # Back-fill planned monthly values for non-actual months so they
                # appear in the Data Entry table and improve the ClosingStock calc.
                for month in MONTHS:
                    if not flags.get(month, False):
                        ratio = float(dist_yr.get(month) or 0)
                        scale = plan_annual if plan_annual > 0 else prev_annual_usd
                        sale_usd_entry.setdefault(yr, {})[month] = scale * ratio

                new_vals = {}
                for month in MONTHS:
                    if flags.get(month, False):
                        sale = float(sale_yr_usd.get(month) or 0)
                    else:
                        ratio = float(dist_yr.get(month) or 0)
                        scale = plan_annual if plan_annual > 0 else prev_annual_usd
                        sale = scale * ratio
                    received = float(planned_yr.get(month) or 0)
                    closing_usd = max(closing_usd - sale + received, 0.0)
                    new_vals[month] = (
                        closing_usd / (annual_sale / 12) if annual_sale > 0 else 0.0
                    )

                sh_usd_entry[yr] = new_vals
                sh_usd_entry[f"total_{yr}"] = (
                    sum(new_vals.values()) / 12
                    if any(v > 0 for v in new_vals.values()) else 0.0
                )

        # ── QTY ────────────────────────────────────────────────────────────────
        received_qty_entry = coll.get("Act. Goods Received QTY")
        sale_qty_entry     = coll.get("Act. Sale QTY")
        sh_qty_entry       = coll.get("Act. Stock Holding Month-QTY")

        if received_qty_entry and sale_qty_entry and sh_qty_entry:
            closing_qty = 0.0
            prev_annual_qty = 0.0

            for yr in ALL_YEARS:
                received_yr  = received_qty_entry.get(yr, {})
                sale_yr_qty  = sale_qty_entry.get(yr, {})

                # Reuse the USD actual flags: months where USD was non-zero are
                # considered "actual" months for QTY too.  This correctly handles
                # a genuine QTY=0 in an actual month (e.g., Mar=0 sold but Mar
                # USD=555 confirms the month had real data).
                flags = usd_actual_flags.get(yr,
                    {m: (float(sale_yr_qty.get(m) or 0) > 0) for m in MONTHS}
                )

                dist_yr = (dist_entry or {}).get(yr, {})

                # For QTY there is no explicit plan-QTY annual total, so scale by
                # prior year's actual QTY total as the mock annual denominator.
                annual_sale = _blended_annual(
                    sale_yr_qty, 0.0, dist_yr, prev_annual_qty, flags
                )

                actual_sum = sum(float(sale_yr_qty.get(m) or 0) for m in MONTHS)
                if actual_sum > 0:
                    prev_annual_qty = actual_sum

                # Back-fill planned monthly QTY for non-actual months
                for month in MONTHS:
                    if not flags.get(month, False):
                        ratio = float(dist_yr.get(month) or 0)
                        sale_qty_entry.setdefault(yr, {})[month] = prev_annual_qty * ratio

                new_vals = {}
                for month in MONTHS:
                    if flags.get(month, False):
                        sale = float(sale_yr_qty.get(month) or 0)
                    else:
                        ratio = float(dist_yr.get(month) or 0)
                        sale = prev_annual_qty * ratio
                    received = float(received_yr.get(month) or 0)
                    closing_qty = max(closing_qty - sale + received, 0.0)
                    new_vals[month] = (
                        closing_qty / (annual_sale / 12) if annual_sale > 0 else 0.0
                    )

                sh_qty_entry[yr] = new_vals
                sh_qty_entry[f"total_{yr}"] = (
                    sum(new_vals.values()) / 12
                    if any(v > 0 for v in new_vals.values()) else 0.0
                )


# ─── DataFrame builder ─────────────────────────────────────────────────────────
def build_dataframe(data: dict, edit_year: str = "2026", ref_year: str = "2025") -> pd.DataFrame:
    """
    Flatten the OTB data dict into a pandas DataFrame with one row per
    (collection × metric) combination.

    Columns:
      _row_idx  _key  _metric
      Collection  Category  Segment  Metric
      Jan'26 ... Dec'26  Total 2026    (editable)
      Jan'25 ... Dec'25  Total 2025    (read-only reference)
    """
    meta = data["_meta"]
    ey, ry = edit_year, ref_year
    ey2, ry2 = ey[2:], ry[2:]

    rows = []
    for key in meta["collections"]:
        if key not in data:
            continue
        collection, category, segment = key

        for metric in EDITABLE_METRICS + DISPLAY_METRICS:
            if metric not in data[key]:
                continue

            is_editable = metric in EDITABLE_METRICS
            entry = data[key][metric]
            row: dict = {
                "_row_idx": entry["row_idx"],
                "_key": repr(key),
                "_metric": metric,
                "_editable": is_editable,
                "Collection": collection,
                "Category": category or "",
                "Segment": segment or "",
                "Metric": METRIC_DISPLAY.get(metric, metric),
            }

            # Year columns (stored for both editable and display metrics)
            for month in MONTHS:
                v = entry[ey].get(month, 0)
                row[f"{month}'{ey2}"] = float(v) if v else 0.0

            row[f"Total {ey}"] = sum(row[f"{m}'{ey2}"] for m in MONTHS)

            for month in MONTHS:
                v = entry[ry].get(month, 0)
                row[f"{month}'{ry2} ◀"] = float(v) if v else 0.0

            row[f"Total {ry} ◀"] = sum(row[f"{m}'{ry2} ◀"] for m in MONTHS)

            rows.append(row)

    return pd.DataFrame(rows) if rows else pd.DataFrame()


# ─── Change detection ──────────────────────────────────────────────────────────
def detect_changes(original_data: dict, edited_df: pd.DataFrame,
                   edit_year: str = "2026") -> list:
    """
    Compare edited_df against original_data to find changed cells.
    Returns a list of change-dicts suitable for save_changes().
    """
    col_map = original_data["_meta"]["col_map"]
    yr = edit_year
    yr2 = yr[2:]
    changes = []

    for _, row in edited_df.iterrows():
        key_str = row.get("_key", "")
        metric = row.get("_metric", "")

        if not key_str or not metric:
            continue

        # Never write back display-only metrics
        if metric in DISPLAY_METRICS:
            continue

        try:
            key = eval(key_str)   # safely reconstruct tuple
        except Exception:
            continue

        if key not in original_data or metric not in original_data[key]:
            continue

        entry = original_data[key][metric]
        row_idx = entry["row_idx"]

        for i, month in enumerate(MONTHS):
            col_name = f"{month}'{yr2}"
            if col_name not in row.index:
                continue

            new_val = row[col_name]
            old_val = entry[yr].get(month, 0) or 0

            # Normalize
            if new_val is None or (isinstance(new_val, float) and pd.isna(new_val)):
                new_val = 0.0
            new_val = float(new_val)
            old_val = float(old_val)

            if abs(new_val - old_val) > 0.0001:
                if i < len(col_map[yr]):
                    changes.append({
                        "row_idx": row_idx,
                        "col_idx": col_map[yr][i],   # 0-based
                        "value": new_val,
                        "metric": METRIC_DISPLAY.get(metric, metric),
                        "month": col_name,
                        "old_value": old_val,
                        "collection": row.get("Collection", ""),
                    })

    return changes


# ─── Validation ────────────────────────────────────────────────────────────────
def validate_changes(changes: list) -> list:
    """
    Return list of error strings.  Empty list = all valid.
    Rules:
      - No negative values
      - QTY metrics must be whole numbers
    """
    errors = []
    for c in changes:
        val = c.get("value")
        metric = c.get("metric", "")
        month = c.get("month", "")
        collection = c.get("collection", "")
        label = f"[{collection} / {metric} / {month}]"

        if val is None:
            continue

        try:
            num = float(val)
        except (ValueError, TypeError):
            errors.append(f"{label} Not a number: {val!r}")
            continue

        if num < 0:
            errors.append(f"{label} Negative values not allowed ({num:,.2f})")

        if "QTY" in metric and num != int(num):
            errors.append(f"{label} QTY must be a whole number, got {num}")

    return errors


# ─── Save ──────────────────────────────────────────────────────────────────────
def save_changes(source, changes: list, dry_run: bool = False) -> dict:
    """
    Write only the changed cells back to Excel, preserving all formulas.
    `source` is either a local file path (str) or raw bytes (cloud mode).

    Returns:
    {
        "success": bool,
        "written": int,
        "errors": [str],
        "backup_path": str | None,
        "dry_run": bool,
        "preview": [str],       (only when dry_run=True)
        "download_bytes": bytes, (only when source is bytes — cloud mode)
    }
    """
    from io import BytesIO as _BytesIO
    cloud_mode = isinstance(source, (bytes, bytearray))

    result: dict = {
        "success": False,
        "written": 0,
        "errors": [],
        "backup_path": None,
        "dry_run": dry_run,
        "preview": [],
    }

    if not changes:
        result["success"] = True
        return result

    if dry_run:
        result["success"] = True
        result["written"] = len(changes)
        result["preview"] = [
            f"  [{c['collection']}] {c['metric']} / {c['month']}: "
            f"{c['old_value']:,.2f} → {c['value']:,.2f}"
            for c in changes
        ]
        return result

    # --- Backup (local only) ---
    if not cloud_mode:
        try:
            result["backup_path"] = backup_file(source)
        except Exception as e:
            result["errors"].append(f"Backup failed: {e}")
            return result

    # --- Write back (no data_only — preserves formulas in other cells) ---
    try:
        buf = _BytesIO(source) if cloud_mode else source
        wb = openpyxl.load_workbook(buf)
        ws = wb[SHEET_NAME]

        for change in changes:
            row_idx = change["row_idx"]         # 1-based
            col_idx = change["col_idx"] + 1     # 0-based → 1-based for openpyxl
            value = change["value"]

            if value is None:
                value = 0
            if "QTY" in change.get("metric", ""):
                value = int(round(value))

            ws.cell(row=row_idx, column=col_idx).value = value
            result["written"] += 1

        if cloud_mode:
            out = _BytesIO()
            wb.save(out)
            wb.close()
            result["download_bytes"] = out.getvalue()
        else:
            wb.save(source)
            wb.close()

        result["success"] = True

    except PermissionError:
        result["errors"].append(
            "⚠️  File is open in Excel.  Close it, then save again."
        )
    except Exception as exc:
        result["errors"].append(str(exc))

    return result


# ─── Backup ────────────────────────────────────────────────────────────────────
def backup_file(filepath: str) -> str:
    """Copy the file to a timestamped backup in the same directory."""
    path = Path(filepath)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = path.parent / f"{path.stem}_backup_{ts}{path.suffix}"
    shutil.copy2(filepath, backup)
    return str(backup)


# ─── KPI aggregation ───────────────────────────────────────────────────────────
def compute_kpis(data: dict, edit_year: str = "2026") -> dict:
    """Aggregate KPI numbers across all collections."""
    meta = data.get("_meta", {})
    total_planned = total_sale_usd = total_qty_received = 0.0
    total_sale_qty = 0.0
    monthly_planned = {m: 0.0 for m in MONTHS}
    planned_by_category: dict = {}

    for key in meta.get("collections", []):
        if key not in data:
            continue
        _, category, _ = key
        cat = category or "Other"

        for metric, entry in data[key].items():
            if metric.startswith("_") or not isinstance(entry, dict):
                continue
            vals = entry.get(edit_year, {})
            total = sum((v or 0) for v in vals.values())

            if metric == "Planned Purchases USD":
                total_planned += total
                for m in MONTHS:
                    monthly_planned[m] += (vals.get(m) or 0)
                planned_by_category[cat] = planned_by_category.get(cat, 0) + total

            elif metric == "Act. Sale USD":
                total_sale_usd += total

            elif metric == "Act. Sale QTY":
                total_sale_qty += total

            elif metric == "Act. Goods Received QTY":
                total_qty_received += total

    return {
        "total_planned":        total_planned,
        "total_sale_usd":       total_sale_usd,
        "total_sale_qty":       total_sale_qty,
        "total_qty_received":   total_qty_received,
        "monthly_planned":      monthly_planned,
        "planned_by_category":  planned_by_category,
    }
